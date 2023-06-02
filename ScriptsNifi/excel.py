from operator import truediv
import pandas as pd
import datetime
from datetime import datetime, timedelta
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.chart import BarChart, Reference
import io 
import sys
import re

def create_excel(lines_df, initial_row_count, warnings_count, rejections_count, file_name,file_type):
    rules_and_significations = {
        "V-Today-1": "<= TODAY",
        "V-DateOfBirth-1": "not > 125 years ago",
        "V-FormatDate-1":"Invalid date format. Must be YYYY-MM-DD HH: MM: SS",
        "V-DateofDeath": "Date must be greater than or equal to DateOfBirth",
        "V-NotNull-1": "<> NULL / blank (Rejet)",
        "V-NotNull-2": "<> NULL / blank (Avertissement)",
        "V-length50": "Longueur ne doit pas dépasser 50 Charactères",
        "V-length100": "Longueur ne doit pas dépasser 100 Charactères",
        "V-alpha-1": "Alpha Characters only",
        "V-alpha-2": "Alpha Characters only",
        "Deduplication" : "Deduplicated lines based on a unique value per line, or the whole line duplicated",
        "Absence MandatoryField" : "A mandatory field was missing",
        "D-Age-1":"Default value applied for Age",
        "D-RoomNumber-1" : "RoomNumber is missing",
        "D-BedNumber-1" : "BedNumber is missing",
        "V-Num-1":"Value must be numeric only",
        "V-Quantity-1":"Value must be greater than zero",
        "D-Duration-1":"Default value applied for Duration",
        "V-GTE0-1":"Value must be greater than or equal to zero"
    }

    # Calculate total warnings
    total_warnings = sum(sum(d.values()) for d in warnings_count.values())

    # Calculate total rejections
    total_rejections = sum(sum(d.values()) for d in rejections_count.values())

    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()

    worksheet = wb.active
    worksheet.title = 'Summary'
    # Merge cells D2:E3 and set the fill color to blue
    cell_range = 'D2:E3'
    worksheet.merge_cells(cell_range)

    fill = PatternFill(start_color='BDD7EE',
                       end_color='BDD7EE', fill_type='solid')
    for row in worksheet[cell_range]:
        for cell in row:
            cell.fill = fill

    # Write the text "Validation report" in black font
    cell = worksheet.cell(row=2, column=4)
    cell.value = "Validation report"
    cell.font = Font(color='000000', bold=True)

    # Center the text horizontally and vertically
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Créer une feuille de calcul pour les avertissements
    ws_warnings = wb.create_sheet('Details')

    # Écrire le DataFrame des avertissements dans la feuille de calcul
    for r in dataframe_to_rows(lines_df, index=False, header=True):
        ws_warnings.append(r)

    # Définir le tableau
    table = [
        'Cluster',
        'Hopital',
        'Servicing department'
    ]

    regex_pattern = r"^([^_]+)"
    match = re.match(regex_pattern, file_name)
    if match:
        cluster_name = match.group(1)
    else:
        cluster_name = ''

    hospital_name = re.search(r'_(.*?)_', file_name)
    if hospital_name:
        hospital_name = hospital_name.group(1)
    else:
        hospital_name = ''

    servicingDepartment = ''
    if file_type == 'Service':
        match = re.search(r'Serv\.([A-Za-z0-9.]+)_', file_name)
        if match:
            servicingDepartment = match.group(1)

    table2 = [cluster_name, hospital_name, servicingDepartment]
    for row in range(4, 7):
        cell = worksheet.cell(row=row, column=4)
        cell.value = table2[row - 4]


    # Écrire les valeurs dans les cellules appropriées
    for row in range(4, 7):
        cell = worksheet.cell(row=row, column=2)
        cell.value = table[row-4]

    # Fusionner les cellules B4:C4, B5:C5, B6:C6
    for row in range(4, 7):
        start_col, start_row, end_col, end_row = range_boundaries(
            f'B{row}:C{row}')
        worksheet.merge_cells(
            start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

    # Mettre le texte en gras pour les titres
    title_font = Font(bold=True)

    for row in range(4, 7):
        cell = worksheet.cell(row=row, column=2)
        cell.font = title_font

    # Définir le remplissage pour le fond en vert kaki
    fill = PatternFill(start_color='C2D69B',
                       end_color='C2D69B', fill_type='solid')

    # Appliquer le remplissage et les bordures noires aux cellules B4:C6
    for row in range(4, 7):
        for col in range(2, 5):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = fill
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin',
                                            color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    # Centrer le texte dans les cellules
    alignment = Alignment(horizontal='center', vertical='center')

    for row in range(4, 7):
        for col in range(2, 5):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = alignment

    # Write "Production date" in bold font in cell D7
    cell = worksheet['D7']
    cell.value = 'Production date'
    cell.font = Font(bold=True)

    # Write today's date in cells F7 to H7
    cell = worksheet['F7']
    cell.value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Appliquer  les bordures noires aux cellules B4:C6
    for col in range(4, 8):
        cell = worksheet.cell(row=7, column=col)
        if col == 5 or col == 7:
            cell.border = Border(
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

        elif col == 4:
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
        else:
            cell.border = Border(
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    # Écrire les valeurs dans les cellules appropriées
    table = [
        'Total number of initial records:',
        'Number of rejected records:',
        'Number of warned records:'
    ]

    for row in range(14, 17):
        cell = worksheet.cell(row=row, column=2)
        cell.value = table[row-14]

    # Appliquer  les bordures noires aux cellules B4:C6
    for row in range(14, 17):
        for col in range(2, 6):
            cell = worksheet.cell(row=row, column=col)
            if col == 3 or col == 4:
                cell.border = Border(
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

            elif col == 2:
                cell.border = Border(left=Side(border_style='thin', color='000000'),
                                     top=Side(border_style='thin',
                                              color='000000'),
                                     bottom=Side(border_style='thin', color='000000'))
            else:
                cell.border = Border(
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Définir les cellules à fusionner
    merge_ranges = ['B14:E14', 'B15:E15', 'B16:E16']

    # Fusionner les cellules
    for cell_range in merge_ranges:
        worksheet.merge_cells(cell_range)

    # Écrire les valeurs dans les cellules appropriées
    table = [
        'Records',
        '%']

    for col in range(6, 8):
        cell = worksheet.cell(row=13, column=col)
        cell.value = table[col-6]
        font = Font(color='0070C0', bold=True)
        cell.font = font
        fill = PatternFill(start_color='BDD7EE',
                           end_color='BDD7EE', fill_type='solid')
        cell.fill = fill
        cell.alignment = alignment
        # Appliquer  les bordures noires aux cellules B4:C6
    for row in range(13, 17):
        for col in range(6, 8):
            cell = worksheet.cell(row=row, column=col)
            cell.border = Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    cell = worksheet.cell(row=14, column=6)
    cell.value = initial_row_count
    cell.alignment = alignment
    cell = worksheet.cell(row=15, column=6)
    cell.value = total_rejections
    cell.alignment = alignment
    cell = worksheet.cell(row=16, column=6)
    cell.value = total_warnings
    cell.alignment = alignment
    cell = worksheet.cell(row=14, column=7)
    cell.value = '100%'
    cell.alignment = alignment
    warning_percentage = (total_warnings / initial_row_count) * 100
    rejection_percentage = (total_rejections / initial_row_count) * 100
    cell = worksheet.cell(row=15, column=7)
    cell.value = str("{:.2f}".format(rejection_percentage))+'%'
    cell.alignment = alignment
    cell = worksheet.cell(row=16, column=7)
    cell.value = str("{:.2f}".format(warning_percentage))+'%'
    cell.alignment = alignment

    table = [
        "Field name",
        "Validation Type",
        "Rule ID",
        "Validation rule",
        "Number",
        "%"
    ]

    for col, value in enumerate(table, start=2):
        cell = worksheet.cell(row=19, column=col)
        cell.value = value
        font = Font(color='0070C0', bold=True)
        cell.font = font
        fill = PatternFill(start_color='BDD7EE',
                           end_color='BDD7EE', fill_type='solid')
        cell.fill = fill
        cell.alignment = alignment

    row = 20

    # Parcourir le dictionnaire warnings_count
    for key, value in warnings_count.items():
        for keyy, valuee in value.items():
            if valuee != 0:
                worksheet.cell(row=row, column=2).value = keyy
                worksheet.cell(row=row, column=3).value = 'Warning'
                worksheet.cell(row=row, column=4).value = key
                worksheet.cell(
                    row=row, column=5).value = rules_and_significations[key]
                worksheet.cell(row=row, column=6).value = valuee
                worksheet.cell(row=row, column=6).alignment = alignment
                worksheet.cell(row=row, column=7).value = str(
                    "{:.2f}".format((valuee / initial_row_count) * 100))+'%'
                worksheet.cell(row=row, column=7).alignment = alignment
                row += 1
                for col in range(2, 8):
                    cell = worksheet.cell(row=row-1, column=col)
                    cell.border = Border(
                        left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
    # Parcourir le dictionnaire rejections_count
    for key, value in rejections_count.items():
        for keyy, valuee in value.items():
            if valuee != 0:
                worksheet.cell(row=row, column=2).value = keyy
                worksheet.cell(row=row, column=3).value = 'Rejection'
                worksheet.cell(row=row, column=4).value = key
                worksheet.cell(
                    row=row, column=5).value = rules_and_significations[key]
                worksheet.cell(row=row, column=6).value = valuee
                worksheet.cell(row=row, column=6).alignment = alignment
                worksheet.cell(row=row, column=7).value = str(
                    "{:.2f}".format((valuee / initial_row_count) * 100))+'%'
                worksheet.cell(row=row, column=7).alignment = alignment
                row += 1
                for col in range(2, 8):
                    cell = worksheet.cell(row=row-1, column=col)
                    cell.border = Border(
                        left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))






    start_row = 20  # Ligne de départ
    end_row = worksheet.max_row

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Distribution du nombre de lignes rejetées/averties"
    chart.y_axis.title = "Nombre de lignes rejetées/averties"
    chart.x_axis.title = "Règles"

    for row_number in range(start_row, end_row + 1):
        # Ajouter une série de données pour chaque ligne
        data_ref = Reference(worksheet, min_col=6, min_row=row_number, max_row=row_number)
        # Add data without using titles from data
        chart.add_data(data_ref, titles_from_data=False)


    # Ajouter le graphique à la feuille de calcul
    chart_location = "H19"  # Emplacement du graphique sur la feuille de calcul
    worksheet.add_chart(chart, chart_location)




    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 60
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['H'].width = 20
    worksheet.column_dimensions['I'].width = 20
    worksheet.column_dimensions['J'].width = 20

    wb.save('/opt/nifi/nifi-current/scripts/results/ValidationReport.xlsx')
